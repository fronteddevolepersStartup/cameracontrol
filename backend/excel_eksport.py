"""
Excel va PDF Eksport Moduli
Professional, rasmiy hujjat formati
"""

import os
from datetime import datetime, date, timedelta
from database import get_connection

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_MAVJUD = True
except ImportError:
    OPENPYXL_MAVJUD = False

EKSPORT_PAPKA = os.path.join(os.path.dirname(__file__), "exports")
os.makedirs(EKSPORT_PAPKA, exist_ok=True)

# ── Ranglar (professional, och fon) ──────────
SARLAVHA_FONI   = "1F4E79"   # To'q ko'k - sarlavha
USTUN_FONI      = "2E75B6"   # Ko'k - ustun
USTUN_MATN      = "FFFFFF"   # Oq matn
KIRDI_FONI      = "E2EFDA"   # Och yashil - kirdi
CHIQDI_FONI     = "FCE4D6"   # Och qizil - chiqdi
JUFT_QATOR      = "F2F7FB"   # Juda och ko'k - juft qator
TOQ_QATOR       = "FFFFFF"   # Oq - toq qator
CHEGARA_RANGI   = "B8CCE4"   # Ko'k-kulrang chegara
SARLAVHA_MATN   = "FFFFFF"   # Oq matn
QORA            = "1A1A1A"   # Qora matn
KULRANG         = "595959"   # Kulrang matn
YASHIL_MATN     = "375623"   # To'q yashil matn
QIZIL_MATN      = "833C00"   # To'q qizil matn
XULOSA_FONI     = "DEEAF1"   # Och ko'k - xulosa


def ing_chegara():
    c = Side(style="thin", color=CHEGARA_RANGI)
    return Border(left=c, right=c, top=c, bottom=c)


def qalin_chegara():
    c = Side(style="medium", color="2E75B6")
    return Border(left=c, right=c, top=c, bottom=c)


def qollay(katak, font=None, fill=None, alignment=None, border=None):
    if font:      katak.font      = font
    if fill:      katak.fill      = fill
    if alignment: katak.alignment = alignment
    if border:    katak.border    = border


# ═══════════════════════════════════════════════════════════
# EXCEL HISOBOT
# ═══════════════════════════════════════════════════════════
def excel_yaratish(davr="bugun"):
    if not OPENPYXL_MAVJUD:
        raise ImportError("openpyxl topilmadi. Terminal: pip install openpyxl")

    bugun = date.today()
    if davr == "bugun":
        bosh = tug = str(bugun)
    elif davr == "hafta":
        bosh = str(bugun - timedelta(days=bugun.weekday()))
        tug  = str(bugun)
    elif davr == "oy":
        bosh = str(bugun.replace(day=1))
        tug  = str(bugun)
    else:
        bosh = "2000-01-01"
        tug  = str(bugun)

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _transport_varaqi(wb, bosh, tug)
    _ishchi_varaqi(wb, bosh, tug)
    _xulosa_varaqi(wb, bosh, tug)

    # Birinchi varaqni faollashtirish
    wb.active = wb["Xulosa"]

    vaqt      = datetime.now().strftime("%Y%m%d_%H%M%S")
    fayl_nomi = "Hisobot_{}_{}.xlsx".format(davr, vaqt)
    yol       = os.path.join(EKSPORT_PAPKA, fayl_nomi)
    wb.save(yol)
    print("[Excel] Saqlandi: {}".format(yol))
    return yol


# ───────────────────────────────────────────────────────────
# XULOSA VARAQI
# ───────────────────────────────────────────────────────────
def _xulosa_varaqi(wb, bosh, tug):
    ws = wb.create_sheet("Xulosa")
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation   = "portrait"
    ws.page_setup.paperSize     = 9   # A4
    ws.page_margins.left        = 0.7
    ws.page_margins.right       = 0.7
    ws.page_margins.top         = 0.75
    ws.page_margins.bottom      = 0.75

    conn = get_connection()
    kirdi_son  = conn.execute("SELECT COUNT(*) FROM transport_log WHERE harakat='kirdi'  AND date(vaqt) BETWEEN ? AND ?", (bosh, tug)).fetchone()[0]
    chiqdi_son = conn.execute("SELECT COUNT(*) FROM transport_log WHERE harakat='chiqdi' AND date(vaqt) BETWEEN ? AND ?", (bosh, tug)).fetchone()[0]
    fura_son   = conn.execute("SELECT COUNT(*) FROM transport_log WHERE tur='Fura'   AND date(vaqt) BETWEEN ? AND ?", (bosh, tug)).fetchone()[0]
    yengil_son = conn.execute("SELECT COUNT(*) FROM transport_log WHERE tur='Yengil' AND date(vaqt) BETWEEN ? AND ?", (bosh, tug)).fetchone()[0]
    ishchi_son = conn.execute("SELECT COUNT(DISTINCT ism) FROM ishchi_log WHERE date(vaqt) BETWEEN ? AND ?", (bosh, tug)).fetchone()[0]
    kunlik     = conn.execute("""
        SELECT date(vaqt) as kun, harakat, COUNT(*) as son
        FROM transport_log WHERE date(vaqt) BETWEEN ? AND ?
        GROUP BY kun, harakat ORDER BY kun
    """, (bosh, tug)).fetchall()
    conn.close()

    # Ustun kengliklari
    kengliklar = {"A": 5, "B": 22, "C": 15, "D": 15, "E": 15, "F": 15}
    for harf, kengl in kengliklar.items():
        ws.column_dimensions[harf].width = kengl

    qator = 1

    # ── Muassasa sarlavhasi ──────────────────
    ws.row_dimensions[qator].height = 14
    ws.merge_cells("A{}:F{}".format(qator, qator))
    k = ws.cell(row=qator, column=1, value="ZAVOD XAVFSIZLIK VA MONITORING TIZIMI")
    qollay(k,
        font=Font(name="Times New Roman", size=14, bold=True, color=SARLAVHA_MATN),
        fill=PatternFill("solid", fgColor=SARLAVHA_FONI),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=qalin_chegara(),
    )
    ws.row_dimensions[qator].height = 28
    qator += 1

    ws.merge_cells("A{}:F{}".format(qator, qator))
    k = ws.cell(row=qator, column=1,
        value="UMUMIY HISOBOT   |   {}  —  {}   |   Tuzildi: {}".format(
            bosh, tug, datetime.now().strftime("%d.%m.%Y  %H:%M")
        ))
    qollay(k,
        font=Font(name="Times New Roman", size=10, color=SARLAVHA_MATN),
        fill=PatternFill("solid", fgColor="2E75B6"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[qator].height = 20
    qator += 1

    # Bo'sh qator
    ws.row_dimensions[qator].height = 8
    qator += 1

    # ── Statistika jadvali ───────────────────
    ws.merge_cells("A{}:F{}".format(qator, qator))
    k = ws.cell(row=qator, column=1, value="1. UMUMIY KO'RSATKICHLAR")
    qollay(k,
        font=Font(name="Times New Roman", size=11, bold=True, color=QORA),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[qator].height = 20
    qator += 1

    stat_data = [
        ("Jami transport kirdi",      kirdi_son,              "ta"),
        ("Jami transport chiqdi",     chiqdi_son,             "ta"),
        ("Shu jumladan furalar",       fura_son,              "ta"),
        ("Shu jumladan yengil mashina", yengil_son,           "ta"),
        ("Ishchi xodimlari (unikal)", ishchi_son,             "nafar"),
        ("Jami transport harakati",   kirdi_son + chiqdi_son, "ta"),
    ]

    for nom, son, birlik in stat_data:
        ws.row_dimensions[qator].height = 20
        # Nom
        k = ws.cell(row=qator, column=2, value=nom)
        qollay(k,
            font=Font(name="Times New Roman", size=10, color=QORA),
            fill=PatternFill("solid", fgColor=JUFT_QATOR if qator % 2 == 0 else TOQ_QATOR),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=ing_chegara(),
        )
        # Son
        k = ws.cell(row=qator, column=3, value=son)
        qollay(k,
            font=Font(name="Times New Roman", size=11, bold=True, color="1F4E79"),
            fill=PatternFill("solid", fgColor=JUFT_QATOR if qator % 2 == 0 else TOQ_QATOR),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=ing_chegara(),
        )
        # Birlik
        k = ws.cell(row=qator, column=4, value=birlik)
        qollay(k,
            font=Font(name="Times New Roman", size=10, color=KULRANG),
            fill=PatternFill("solid", fgColor=JUFT_QATOR if qator % 2 == 0 else TOQ_QATOR),
            alignment=Alignment(horizontal="left", vertical="center"),
            border=ing_chegara(),
        )
        qator += 1

    ws.row_dimensions[qator].height = 10
    qator += 1

    # ── Kunlik jadval ─────────────────────────
    ws.merge_cells("A{}:F{}".format(qator, qator))
    k = ws.cell(row=qator, column=1, value="2. KUNLIK TRANSPORT HARAKATI")
    qollay(k,
        font=Font(name="Times New Roman", size=11, bold=True, color=QORA),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[qator].height = 20
    qator += 1

    # Ustun sarlavha
    ws.row_dimensions[qator].height = 22
    for col_i, (nom, kengl) in enumerate([
        ("Sana", 22), ("Kirdi (ta)", 15), ("Chiqdi (ta)", 15), ("Jami (ta)", 15)
    ], 2):
        k = ws.cell(row=qator, column=col_i, value=nom)
        qollay(k,
            font=Font(name="Times New Roman", size=10, bold=True, color=USTUN_MATN),
            fill=PatternFill("solid", fgColor=USTUN_FONI),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=ing_chegara(),
        )
    qator += 1

    # Kunlik ma'lumotlar
    kunlar = {}
    for q in kunlik:
        k2 = q["kun"]
        if k2 not in kunlar:
            kunlar[k2] = {"kirdi": 0, "chiqdi": 0}
        kunlar[k2][q["harakat"]] = q["son"]

    for j, (kun, h) in enumerate(sorted(kunlar.items())):
        ws.row_dimensions[qator].height = 18
        bg = JUFT_QATOR if j % 2 == 0 else TOQ_QATOR
        jami = h["kirdi"] + h["chiqdi"]
        for col_i, qiy in enumerate([kun, h["kirdi"], h["chiqdi"], jami], 2):
            k = ws.cell(row=qator, column=col_i, value=qiy)
            qollay(k,
                font=Font(name="Times New Roman", size=10,
                          bold=(col_i == 5), color=QORA),
                fill=PatternFill("solid", fgColor=bg),
                alignment=Alignment(horizontal="center" if col_i > 2 else "left",
                                    vertical="center"),
                border=ing_chegara(),
            )
        qator += 1

    ws.row_dimensions[qator].height = 16
    qator += 1

    # ── Imzo qismi ───────────────────────────
    ws.row_dimensions[qator].height = 18
    ws.merge_cells("A{}:F{}".format(qator, qator))
    k = ws.cell(row=qator, column=1,
        value="Hisobotni tuzdi: Xavfsizlik xizmati    "
              "Sana: ____________    Imzo: ____________")
    qollay(k,
        font=Font(name="Times New Roman", size=10, color=KULRANG, italic=True),
        alignment=Alignment(horizontal="right", vertical="center"),
    )

    return ws


# ───────────────────────────────────────────────────────────
# TRANSPORT VARAQI
# ───────────────────────────────────────────────────────────
def _transport_varaqi(wb, bosh, tug):
    ws = wb.create_sheet("Transport Jurnali")
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.paperSize     = 9
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1
    ws.page_margins.left        = 0.5
    ws.page_margins.right       = 0.5
    ws.page_margins.top         = 0.75
    ws.page_margins.bottom      = 0.75
    ws.freeze_panes             = "A5"

    conn = get_connection()
    qatorlar = conn.execute("""
        SELECT * FROM transport_log
        WHERE date(vaqt) BETWEEN ? AND ?
        ORDER BY vaqt DESC
    """, (bosh, tug)).fetchall()
    conn.close()

    # Ustun kengliklari
    for harf, kengl in zip("ABCDEFGHI", [5, 20, 14, 10, 10, 14, 18, 11, 16]):
        ws.column_dimensions[harf].width = kengl

    # ── Sarlavha ─────────────────────────────
    ws.merge_cells("A1:I1")
    k = ws["A1"]
    k.value = "TRANSPORT KIRISH-CHIQISH JURNALI"
    qollay(k,
        font=Font(name="Times New Roman", size=14, bold=True, color=SARLAVHA_MATN),
        fill=PatternFill("solid", fgColor=SARLAVHA_FONI),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=qalin_chegara(),
    )
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    k = ws["A2"]
    k.value = "Davr: {}  —  {}          Jami: {} ta yozuv          Tuzildi: {}".format(
        bosh, tug, len(qatorlar), datetime.now().strftime("%d.%m.%Y %H:%M")
    )
    qollay(k,
        font=Font(name="Times New Roman", size=10, color=SARLAVHA_MATN),
        fill=PatternFill("solid", fgColor="2E75B6"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6

    # ── Ustun sarlavhalar ────────────────────
    ws.row_dimensions[4].height = 24
    ustunlar = ["№", "Vaqt", "Davlat raqami", "Turi", "Rangi",
                "Davlat", "Viloyat", "Harakati", "Rasm"]
    for col_i, nom in enumerate(ustunlar, 1):
        k = ws.cell(row=4, column=col_i, value=nom)
        qollay(k,
            font=Font(name="Times New Roman", size=10, bold=True, color=USTUN_MATN),
            fill=PatternFill("solid", fgColor=USTUN_FONI),
            alignment=Alignment(horizontal="center", vertical="center",
                                wrap_text=True),
            border=ing_chegara(),
        )

    # ── Ma'lumotlar ───────────────────────────
    for i, q in enumerate(qatorlar, 1):
        qator_num = i + 4
        juft = (i % 2 == 0)
        bg   = JUFT_QATOR if juft else TOQ_QATOR
        ws.row_dimensions[qator_num].height = 52

        harakat = (q["harakat"] or "").lower()
        if harakat == "kirdi":
            h_bg  = KIRDI_FONI
            h_fg  = YASHIL_MATN
            h_txt = "KIRDI"
        elif harakat == "chiqdi":
            h_bg  = CHIQDI_FONI
            h_fg  = QIZIL_MATN
            h_txt = "CHIQDI"
        else:
            h_bg  = bg
            h_fg  = QORA
            h_txt = harakat.upper()

        qiymatlar = [
            (i,             bg,   QORA,     "center", False),
            (q["vaqt"],     bg,   KULRANG,  "center", False),
            (q["raqam"],    bg,   QORA,     "center", True),
            (q["tur"],      bg,   QORA,     "center", False),
            (q["rang"],     bg,   QORA,     "center", False),
            (q["davlat"],   bg,   QORA,     "center", False),
            (q["viloyat"],  bg,   QORA,     "left",   False),
            (h_txt,         h_bg, h_fg,     "center", True),
        ]

        for col_i, (qiy, bg_c, fg_c, align, bold) in enumerate(qiymatlar, 1):
            k = ws.cell(row=qator_num, column=col_i, value=qiy)
            qollay(k,
                font=Font(name="Times New Roman", size=10,
                          bold=bold, color=fg_c),
                fill=PatternFill("solid", fgColor=bg_c),
                alignment=Alignment(horizontal=align, vertical="center"),
                border=ing_chegara(),
            )

        # Rasm
        rasm_yol = q["rasm_yol"] or ""
        if rasm_yol and os.path.exists(rasm_yol):
            try:
                img = XLImage(rasm_yol)
                img.width  = 85
                img.height = 48
                hujayra = "I{}".format(qator_num)
                ws.add_image(img, hujayra)
            except Exception:
                pass
        k = ws.cell(row=qator_num, column=9, value="" if (rasm_yol and os.path.exists(rasm_yol)) else "—")
        qollay(k,
            font=Font(name="Times New Roman", size=9, color=KULRANG),
            fill=PatternFill("solid", fgColor=bg),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=ing_chegara(),
        )

    # ── Xulosa qatori ─────────────────────────
    jami_qator = len(qatorlar) + 5
    ws.row_dimensions[jami_qator].height = 20
    ws.merge_cells("A{}:G{}".format(jami_qator, jami_qator))
    k = ws.cell(row=jami_qator, column=1, value="JAMI:")
    qollay(k,
        font=Font(name="Times New Roman", size=10, bold=True, color=QORA),
        fill=PatternFill("solid", fgColor=XULOSA_FONI),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=ing_chegara(),
    )
    k = ws.cell(row=jami_qator, column=8,
        value="{} ta".format(len(qatorlar)))
    qollay(k,
        font=Font(name="Times New Roman", size=10, bold=True, color="1F4E79"),
        fill=PatternFill("solid", fgColor=XULOSA_FONI),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=ing_chegara(),
    )

    return ws


# ───────────────────────────────────────────────────────────
# ISHCHILAR VARAQI
# ───────────────────────────────────────────────────────────
def _ishchi_varaqi(wb, bosh, tug):
    ws = wb.create_sheet("Ishchilar Jurnali")
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation   = "portrait"
    ws.page_setup.paperSize     = 9
    ws.page_margins.left        = 0.7
    ws.page_margins.right       = 0.7
    ws.freeze_panes             = "A5"

    conn = get_connection()
    qatorlar = conn.execute("""
        SELECT * FROM ishchi_log
        WHERE date(vaqt) BETWEEN ? AND ?
        ORDER BY vaqt DESC
    """, (bosh, tug)).fetchall()
    davomiylik = conn.execute("""
        SELECT ism, COUNT(*) as son, MIN(vaqt) as birinchi, MAX(vaqt) as oxirgi
        FROM ishchi_log WHERE date(vaqt) BETWEEN ? AND ?
        GROUP BY ism ORDER BY son DESC
    """, (bosh, tug)).fetchall()
    conn.close()

    for harf, kengl in zip("ABCDE", [5, 22, 24, 14, 22]):
        ws.column_dimensions[harf].width = kengl

    # Sarlavha
    ws.merge_cells("A1:E1")
    k = ws["A1"]
    k.value = "ISHCHILAR DAVOMIYLIK JURNALI"
    qollay(k,
        font=Font(name="Times New Roman", size=14, bold=True, color=SARLAVHA_MATN),
        fill=PatternFill("solid", fgColor=SARLAVHA_FONI),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=qalin_chegara(),
    )
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    k = ws["A2"]
    k.value = "Davr: {}  —  {}     Jami qayd: {}     Tuzildi: {}".format(
        bosh, tug, len(qatorlar), datetime.now().strftime("%d.%m.%Y %H:%M")
    )
    qollay(k,
        font=Font(name="Times New Roman", size=10, color=SARLAVHA_MATN),
        fill=PatternFill("solid", fgColor="2E75B6"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    # Ustunlar
    ws.row_dimensions[4].height = 22
    for col_i, nom in enumerate(["№", "Vaqt", "Xodim ismi", "Harakat", "Identifikator"], 1):
        k = ws.cell(row=4, column=col_i, value=nom)
        qollay(k,
            font=Font(name="Times New Roman", size=10, bold=True, color=USTUN_MATN),
            fill=PatternFill("solid", fgColor=USTUN_FONI),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=ing_chegara(),
        )

    for i, q in enumerate(qatorlar, 1):
        qn = i + 4
        juft = (i % 2 == 0)
        bg   = JUFT_QATOR if juft else TOQ_QATOR
        ws.row_dimensions[qn].height = 18

        for col_i, qiy in enumerate([i, q["vaqt"], q["ism"], q["harakat"], q["yuz_id"]], 1):
            k = ws.cell(row=qn, column=col_i, value=qiy)
            qollay(k,
                font=Font(name="Times New Roman", size=10,
                          bold=(col_i == 3), color=QORA),
                fill=PatternFill("solid", fgColor=bg),
                alignment=Alignment(
                    horizontal="center" if col_i in (1, 2, 4) else "left",
                    vertical="center"
                ),
                border=ing_chegara(),
            )

    # Davomiylik jadvali
    if davomiylik:
        bosh_d = len(qatorlar) + 7
        ws.row_dimensions[bosh_d - 1].height = 12
        ws.merge_cells("A{}:E{}".format(bosh_d, bosh_d))
        k = ws.cell(row=bosh_d, column=1, value="ISHCHILAR DAVOMIYLIK JADVALI")
        qollay(k,
            font=Font(name="Times New Roman", size=11, bold=True, color=QORA),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        ws.row_dimensions[bosh_d].height = 20

        ws.row_dimensions[bosh_d + 1].height = 22
        for col_i, nom in enumerate(
            ["№", "Xodim ismi", "Qaydlar soni", "Birinchi qayd", "Oxirgi qayd"], 1
        ):
            k = ws.cell(row=bosh_d + 1, column=col_i, value=nom)
            qollay(k,
                font=Font(name="Times New Roman", size=10, bold=True, color=USTUN_MATN),
                fill=PatternFill("solid", fgColor=USTUN_FONI),
                alignment=Alignment(horizontal="center", vertical="center"),
                border=ing_chegara(),
            )

        for j, d in enumerate(davomiylik):
            r  = bosh_d + 2 + j
            bg = JUFT_QATOR if j % 2 == 0 else TOQ_QATOR
            ws.row_dimensions[r].height = 18
            for col_i, qiy in enumerate(
                [j + 1, d["ism"], d["son"], d["birinchi"], d["oxirgi"]], 1
            ):
                k = ws.cell(row=r, column=col_i, value=qiy)
                qollay(k,
                    font=Font(name="Times New Roman", size=10,
                              bold=(col_i == 3), color=QORA),
                    fill=PatternFill("solid", fgColor=bg),
                    alignment=Alignment(
                        horizontal="center" if col_i in (1, 3) else "left",
                        vertical="center"
                    ),
                    border=ing_chegara(),
                )

    return ws


# ═══════════════════════════════════════════════════════════
# PDF HISOBOT (ReportLab orqali)
# ═══════════════════════════════════════════════════════════
def pdf_yaratish(davr="bugun"):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import io
    except ImportError:
        raise ImportError("reportlab topilmadi. Terminal: pip install reportlab")

    bugun = date.today()
    if davr == "bugun":
        bosh = tug = str(bugun)
    elif davr == "hafta":
        bosh = str(bugun - timedelta(days=bugun.weekday()))
        tug  = str(bugun)
    elif davr == "oy":
        bosh = str(bugun.replace(day=1))
        tug  = str(bugun)
    else:
        bosh = "2000-01-01"
        tug  = str(bugun)

    conn = get_connection()
    t_qatorlar = conn.execute("""
        SELECT vaqt, raqam, tur, rang, davlat, viloyat, harakat
        FROM transport_log WHERE date(vaqt) BETWEEN ? AND ?
        ORDER BY vaqt DESC
    """, (bosh, tug)).fetchall()
    i_qatorlar = conn.execute("""
        SELECT vaqt, ism, harakat FROM ishchi_log
        WHERE date(vaqt) BETWEEN ? AND ?
        ORDER BY vaqt DESC
    """, (bosh, tug)).fetchall()
    kirdi_son  = conn.execute("SELECT COUNT(*) FROM transport_log WHERE harakat='kirdi'  AND date(vaqt) BETWEEN ? AND ?", (bosh, tug)).fetchone()[0]
    chiqdi_son = conn.execute("SELECT COUNT(*) FROM transport_log WHERE harakat='chiqdi' AND date(vaqt) BETWEEN ? AND ?", (bosh, tug)).fetchone()[0]
    conn.close()

    vaqt      = datetime.now().strftime("%Y%m%d_%H%M%S")
    fayl_nomi = "Hisobot_{}_{}.pdf".format(davr, vaqt)
    yol       = os.path.join(EKSPORT_PAPKA, fayl_nomi)

    doc = SimpleDocTemplate(
        yol, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm,  bottomMargin=2*cm,
        title="Zavod Monitoring Hisoboti",
    )

    # Ranglar
    KOK      = colors.HexColor("#1F4E79")
    USTUN_BG = colors.HexColor("#2E75B6")
    YASH_BG  = colors.HexColor("#E2EFDA")
    QIZ_BG   = colors.HexColor("#FCE4D6")
    JUFT_BG  = colors.HexColor("#F2F7FB")
    CHEGARA  = colors.HexColor("#B8CCE4")
    OQ_C     = colors.white
    QORA_C   = colors.HexColor("#1A1A1A")

    styles = getSampleStyleSheet()

    def p_stil(size=10, bold=False, color=QORA_C, align="LEFT"):
        return ParagraphStyle(
            "custom",
            fontSize=size,
            fontName="Helvetica-Bold" if bold else "Helvetica",
            textColor=color,
            alignment={"LEFT": 0, "CENTER": 1, "RIGHT": 2}[align],
            leading=size * 1.3,
        )

    taqdimot = []

    # ── Sarlavha ─────────────────────────────
    taqdimot.append(Paragraph(
        "ZAVOD XAVFSIZLIK VA MONITORING TIZIMI",
        p_stil(16, bold=True, color=OQ_C, align="CENTER")
    ))

    sarlavha_jadval = Table(
        [["ZAVOD XAVFSIZLIK VA MONITORING TIZIMI"]],
        colWidths=[17*cm]
    )
    sarlavha_jadval.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,-1), KOK),
        ("TEXTCOLOR",   (0,0), (-1,-1), OQ_C),
        ("FONTNAME",    (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 14),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [KOK]),
        ("TOPPADDING",  (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))
    taqdimot.append(sarlavha_jadval)

    taqdimot.append(Spacer(1, 0.2*cm))

    # Kichik sarlavha
    info_jadval = Table(
        [["Davr: {}  —  {}          Tuzildi: {}".format(
            bosh, tug, datetime.now().strftime("%d.%m.%Y  %H:%M")
        )]],
        colWidths=[17*cm]
    )
    info_jadval.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,-1), USTUN_BG),
        ("TEXTCOLOR",   (0,0), (-1,-1), OQ_C),
        ("FONTNAME",    (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    taqdimot.append(info_jadval)
    taqdimot.append(Spacer(1, 0.4*cm))

    # ── Statistika ───────────────────────────
    taqdimot.append(Paragraph("1. UMUMIY KO'RSATKICHLAR",
                               p_stil(11, bold=True, color=KOK)))
    taqdimot.append(Spacer(1, 0.2*cm))

    stat_data = [
        ["Ko'rsatkich", "Miqdor", "Birlik"],
        ["Jami transport kirdi",       str(kirdi_son),              "ta"],
        ["Jami transport chiqdi",      str(chiqdi_son),             "ta"],
        ["Jami transport harakati",    str(kirdi_son + chiqdi_son), "ta"],
        ["Ishchi xodimlari (unikal)",  str(conn.execute("SELECT COUNT(DISTINCT ism) FROM ishchi_log WHERE date(vaqt) BETWEEN ? AND ?", (bosh, tug)).fetchone()[0] if False else "—"), "nafar"],
    ]

    # Ishchi sonini qayta olish
    conn2 = get_connection()
    ishchi_son = conn2.execute("SELECT COUNT(DISTINCT ism) FROM ishchi_log WHERE date(vaqt) BETWEEN ? AND ?", (bosh, tug)).fetchone()[0]
    conn2.close()
    stat_data[4][1] = str(ishchi_son)

    stat_jadval = Table(stat_data, colWidths=[10*cm, 4*cm, 3*cm])
    stat_jadval.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0),  USTUN_BG),
        ("TEXTCOLOR",    (0,0), (-1,0),  OQ_C),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 10),
        ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
        ("ALIGN",        (1,0), (-1,-1), "CENTER"),
        ("ALIGN",        (0,0), (0,-1),  "LEFT"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [OQ_C, JUFT_BG]),
        ("GRID",         (0,0), (-1,-1), 0.5, CHEGARA),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ("LEFTPADDING",  (0,0), (-1,-1), 8),
        ("FONTNAME",     (1,1), (1,-1),  "Helvetica-Bold"),
        ("TEXTCOLOR",    (1,1), (1,-1),  KOK),
    ]))
    taqdimot.append(stat_jadval)
    taqdimot.append(Spacer(1, 0.4*cm))

    # ── Transport jurnali ────────────────────
    taqdimot.append(Paragraph("2. TRANSPORT KIRISH-CHIQISH JURNALI",
                               p_stil(11, bold=True, color=KOK)))
    taqdimot.append(Spacer(1, 0.2*cm))

    t_sarlavha = [["№", "Vaqt", "Raqam", "Turi", "Rangi", "Davlat/Viloyat", "Harakati"]]
    t_qator_data = []
    for i, q in enumerate(t_qatorlar[:100], 1):  # max 100 ta
        t_qator_data.append([
            str(i),
            str(q["vaqt"])[:16] if q["vaqt"] else "",
            str(q["raqam"] or ""),
            str(q["tur"] or ""),
            str(q["rang"] or ""),
            "{}/{}".format(q["davlat"] or "", q["viloyat"] or ""),
            str(q["harakat"] or "").upper(),
        ])

    t_jadval_data = t_sarlavha + t_qator_data
    t_jadval = Table(
        t_jadval_data,
        colWidths=[1*cm, 3.2*cm, 2.5*cm, 1.8*cm, 1.8*cm, 4*cm, 2*cm],
        repeatRows=1
    )

    t_stil = [
        ("BACKGROUND",   (0,0), (-1,0),  USTUN_BG),
        ("TEXTCOLOR",    (0,0), (-1,0),  OQ_C),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("ALIGN",        (5,1), (5,-1),  "LEFT"),
        ("GRID",         (0,0), (-1,-1), 0.5, CHEGARA),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [OQ_C, JUFT_BG]),
        ("TOPPADDING",   (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0), (-1,-1), 4),
    ]

    # Kirdi/Chiqdi ranglash
    for i, q in enumerate(t_qatorlar[:100], 1):
        if q["harakat"] == "kirdi":
            t_stil.append(("BACKGROUND", (6, i), (6, i), YASH_BG))
            t_stil.append(("TEXTCOLOR",  (6, i), (6, i), colors.HexColor("#375623")))
        elif q["harakat"] == "chiqdi":
            t_stil.append(("BACKGROUND", (6, i), (6, i), QIZ_BG))
            t_stil.append(("TEXTCOLOR",  (6, i), (6, i), colors.HexColor("#833C00")))

    t_jadval.setStyle(TableStyle(t_stil))
    taqdimot.append(t_jadval)
    taqdimot.append(Spacer(1, 0.4*cm))

    # ── Ishchilar jurnali ────────────────────
    if i_qatorlar:
        taqdimot.append(Paragraph("3. ISHCHILAR DAVOMIYLIK JURNALI",
                                   p_stil(11, bold=True, color=KOK)))
        taqdimot.append(Spacer(1, 0.2*cm))

        i_sarlavha = [["№", "Vaqt", "Xodim ismi", "Harakat"]]
        i_qator_data = []
        for i, q in enumerate(i_qatorlar[:100], 1):
            i_qator_data.append([
                str(i),
                str(q["vaqt"])[:16] if q["vaqt"] else "",
                str(q["ism"] or ""),
                str(q["harakat"] or ""),
            ])

        i_jadval_data = i_sarlavha + i_qator_data
        i_jadval = Table(
            i_jadval_data,
            colWidths=[1*cm, 4*cm, 7*cm, 5*cm],
            repeatRows=1
        )
        i_jadval.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0),  USTUN_BG),
            ("TEXTCOLOR",    (0,0), (-1,0),  OQ_C),
            ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
            ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
            ("ALIGN",        (0,0), (-1,-1), "CENTER"),
            ("ALIGN",        (2,1), (2,-1),  "LEFT"),
            ("GRID",         (0,0), (-1,-1), 0.5, CHEGARA),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [OQ_C, JUFT_BG]),
            ("TOPPADDING",   (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
        ]))
        taqdimot.append(i_jadval)
        taqdimot.append(Spacer(1, 0.4*cm))

    # ── Imzo ─────────────────────────────────
    taqdimot.append(HRFlowable(width="100%", thickness=1,
                                color=CHEGARA, spaceAfter=0.2*cm))
    taqdimot.append(Paragraph(
        "Hisobotni tuzdi: Xavfsizlik xizmati &nbsp;&nbsp;&nbsp;&nbsp; "
        "Sana: ____________ &nbsp;&nbsp;&nbsp;&nbsp; Imzo: ____________",
        p_stil(9, color=colors.HexColor("#595959"), align="RIGHT")
    ))

    doc.build(taqdimot)
    print("[PDF] Saqlandi: {}".format(yol))
    return yol
